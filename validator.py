import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import PatternFill
import io
from datetime import datetime, date
import copy

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Academic Personnel Validator",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# STYLING
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
}

/* Sidebar */
section[data-testid="stSidebar"] {
    background: #0f0f0f;
    border-right: 1px solid #2a2a2a;
}
section[data-testid="stSidebar"] * {
    color: #e0e0e0 !important;
}
section[data-testid="stSidebar"] .stFileUploader label {
    color: #aaa !important;
}

/* Main background */
.main .block-container {
    padding-top: 2rem;
    max-width: 1400px;
}

/* Title */
.app-title {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 2rem;
    font-weight: 600;
    color: inherit;
    letter-spacing: -1px;
    border-bottom: 3px solid currentColor;
    padding-bottom: 0.4rem;
    margin-bottom: 0.2rem;
}
.app-subtitle {
    color: #999;
    font-size: 0.9rem;
    margin-bottom: 2rem;
    font-family: 'IBM Plex Mono', monospace;
}

/* Metric cards */
.metric-row {
    display: flex;
    gap: 1rem;
    margin-bottom: 2rem;
    flex-wrap: wrap;
}
.metric-card {
    flex: 1;
    min-width: 140px;
    background: #0f0f0f;
    color: #fff;
    border-radius: 8px;
    padding: 1.2rem 1.4rem;
    display: flex;
    flex-direction: column;
}
.metric-card.green  { background: #1a3a2a; border-left: 4px solid #2ecc71; }
.metric-card.red    { background: #3a1a1a; border-left: 4px solid #e74c3c; }
.metric-card.blue   { background: #1a2a3a; border-left: 4px solid #3498db; }
.metric-card.yellow { background: #3a3a1a; border-left: 4px solid #f1c40f; }
.metric-num {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 2.2rem;
    font-weight: 600;
    line-height: 1;
    color: #fff;
}
.metric-label {
    font-size: 0.75rem;
    color: #aaa;
    margin-top: 0.4rem;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}

/* Section headers */
.section-header {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.85rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: #aaa;
    border-bottom: 1px solid #444;
    padding-bottom: 0.4rem;
    margin: 2rem 0 1rem 0;
}

/* Checkbox grid */
.col-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 0.3rem 1rem;
    background: rgba(255,255,255,0.05);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 8px;
    padding: 1rem 1.2rem;
    margin-bottom: 1rem;
}

/* Error badge */
.err-badge {
    display: inline-block;
    background: #e74c3c;
    color: white;
    border-radius: 999px;
    padding: 0.1rem 0.5rem;
    font-size: 0.72rem;
    font-family: 'IBM Plex Mono', monospace;
    margin-left: 0.4rem;
}

/* Table styling */
.val-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 0.82rem;
    font-family: 'IBM Plex Mono', monospace;
    margin-top: 1rem;
    color: #e0e0e0;
}
.val-table th {
    background: #1a1a2e;
    color: #e0e0e0;
    padding: 0.5rem 0.7rem;
    text-align: left;
    white-space: nowrap;
    position: sticky;
    top: 0;
    z-index: 1;
}
.val-table td {
    padding: 0.4rem 0.7rem;
    border-bottom: 1px solid #2a2a3a;
    white-space: nowrap;
    max-width: 160px;
    overflow: hidden;
    text-overflow: ellipsis;
    color: #d0d0d0;
    background: #16161e;
}
.val-table tr:hover td { background: #22223a !important; }
.cell-ok  { background: #0d2818; color: #4ade80; }
.cell-err { background: #2d1215; color: #f87171; font-weight: 600; }
.cell-skip { background: #1a1a22; color: #666; }
.cell-edited { background: #1a2a3a; color: #60a5fa; border-left: 3px solid #3b82f6; }
.cell-edited-ok { background: #0d2830; color: #38bdf8; border-left: 3px solid #3b82f6; }

/* Edit badge */
.edit-badge {
    display: inline-block;
    background: #3b82f6;
    color: white;
    border-radius: 999px;
    padding: 0.1rem 0.5rem;
    font-size: 0.68rem;
    font-family: 'IBM Plex Mono', monospace;
    margin-left: 0.3rem;
}

/* Run button */
div[data-testid="stButton"] > button {
    background: #1a1a2e;
    color: #e0e0e0;
    border: 1px solid #333;
    border-radius: 6px;
    padding: 0.6rem 2rem;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.9rem;
    font-weight: 600;
    cursor: pointer;
    transition: background 0.2s;
}
div[data-testid="stButton"] > button:hover {
    background: #2a2a4e;
}

/* Error detail rows */
.err-row-header {
    font-family: 'IBM Plex Mono', monospace;
    font-weight: 600;
    font-size: 0.85rem;
    color: #f87171;
}
.err-item {
    font-size: 0.82rem;
    color: #ccc;
    margin: 0.15rem 0 0.15rem 1rem;
    font-family: 'IBM Plex Sans', sans-serif;
}
.err-col-tag {
    display: inline-block;
    background: #2a2a3a;
    color: #e0e0e0;
    border-radius: 4px;
    padding: 0.05rem 0.4rem;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.75rem;
    margin-right: 0.4rem;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# VALIDATION RULES
# ─────────────────────────────────────────────
COLUMNS = {
    "A": "Name",
    "B": "Fathers Name",
    "C": "Grand fathers Name",
    "D": "Gender",
    "E": "Age",
    "F": "Education Level",
    "G": "Field Of Study",
    "H": "University/College",
    "I": "Subject of Teaching",
    "J": "Job Title",
    "K": "Has Taken Course (PGDSL/PGDSLM)",
    "L": "Date of Employment",
    "M": "Career Ladder",
    "N": "School Ownership",
    "O": "School Name",
    "P": "Level Of The School",
    "Q": "SUB-CITY",
    "R": "Type of Licence Owned",
    "S": "Date of Licence Owned",
    "T": "Mobile Number",
    "U": "Disability",
    "V": "Activity Type",
    "W": "Total Experience (Years)",
    "X": "Carrier Number",
}

DROPDOWNS = {
    "D": ["F", "M"],
    "F": ["Diploma", "Bachelor", "Master", "Director", "Doctorate", "Professional"],
    "I": [
        "Management", "Math", "Chemistry", "Physics", "English", "Afan Oromo",
        "Amharic", "HPE", "Geography", "Civics / Citizenship", "History", "Biology",
        "ICT", "General Science", "SNE", "Social Studies", "Environmental Science",
        "Moral Education", "PVA", "HPE / In Amharic", "HPE / In Afan Oromo",
        "Math / In Amharic", "Math / In Afan Oromo", "Environmental / In Amharic",
        "Environmental / In Afan Oromo", "Moral / In Amharic", "Moral / In Afan Oromo",
        "PVA / In Amharic", "PVA / In Afan Oromo", "Principal / In Amharic",
        "Principal / In Afan Oromo", "Supervisor / In Amharic", "Supervisor / In Afan Oromo",
    ],
    "K": ["Yes", "No"],
    "M": ["Beginner", "Junior", "Higher", "Associate", "Associate Lead", "Lead",
          "Senior Lead", "Senior Lead Two", "Senior Lead Three"],
    "N": ["Private", "Public", "Government", "Unknown"],
    "P": ["KG", "Primary School", "Primary and Middle School", "Secondary School"],
    "Q": ["Bole", "Arada", "Gulale", "Lami Kura", "Kolfe Karanio", "Nifas Silk Lafto",
          "Akaki Qality", "Kirkos", "Adis Ketama", "Lideta", "Yeka"],
    "R": ["Temporary", "Entry level", "Full", "Permanent", "SAT"],
    "U": ["None", "HearingImpairment", "VisualImpairment", "MobilityImpairment", "Autism", "Other"],
    "V": ["Teacher", "Director", "ViceDirector", "Supervisor", "Unknown"],
}

REQUIRED = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "T", "U", "V"]
OPTIONAL = ["R", "S", "W", "X"]
DATE_COLS = ["L", "S"]
COL_LETTERS = list(COLUMNS.keys())

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def col_index(letter):
    return ord(letter) - ord("A")

def parse_excel_date(val):
    if val is None or val == "":
        return None, None
    if isinstance(val, (datetime, date)):
        return val, str(val.date() if isinstance(val, datetime) else val)
    try:
        n = float(val)
        dt = from_excel(n)
        return dt, str(dt.date())
    except Exception:
        return None, str(val)

def validate_cell(col_letter, raw_val):
    """Returns (display_value, error_message_or_None)"""
    val = str(raw_val).strip() if raw_val is not None else ""
    is_empty = val == "" or val.lower() == "none"

    # Required check
    if col_letter in REQUIRED and is_empty:
        return val, f"Required field is empty"

    # Skip optional empty fields
    if col_letter in OPTIONAL and is_empty:
        return val, None

    # Dropdown check
    if col_letter in DROPDOWNS and not is_empty:
        allowed = DROPDOWNS[col_letter]
        if val not in allowed:
            return val, f"'{val}' not allowed. Must be one of: {', '.join(allowed)}"

    # Age check
    if col_letter == "E" and not is_empty:
        try:
            age = int(float(val))
            if not (18 <= age <= 80):
                return val, f"Age {age} out of range (18–80)"
        except ValueError:
            return val, f"'{val}' is not a valid number"

    # Date checks
    if col_letter in DATE_COLS and not is_empty:
        dt, display = parse_excel_date(raw_val)
        if dt is None:
            return val, f"'{val}' is not a valid date"
        return display, None

    # Mobile number: 9 digits
    if col_letter == "T" and not is_empty:
        digits = val.replace(" ", "")
        if not digits.isdigit() or len(digits) != 9:
            return val, f"'{val}' must be exactly 9 digits"

    # Total experience: non-negative number
    if col_letter == "W" and not is_empty:
        try:
            exp = float(val)
            if exp < 0:
                return val, f"Experience cannot be negative"
        except ValueError:
            return val, f"'{val}' is not a valid number"

    # Date display formatting
    if col_letter in DATE_COLS and not is_empty:
        _, display = parse_excel_date(raw_val)
        return display, None

    return val, None


def get_sheet_names(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    return wb.sheetnames


def is_row_empty(row):
    """Return True if every cell in the row is None or blank."""
    return all(cell is None or str(cell).strip() == "" for cell in row)


def read_excel_rows(file_bytes, sheet_name=None):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not is_row_empty(row):
            rows.append(row)
    return rows


def rows_to_dataframe(rows):
    """Convert raw row tuples to a DataFrame with column letter headers."""
    data = []
    for row in rows:
        row_dict = {}
        for col_letter in COL_LETTERS:
            idx = col_index(col_letter)
            raw = row[idx] if idx < len(row) else None
            val = str(raw).strip() if raw is not None else ""
            row_dict[col_letter] = val
        data.append(row_dict)
    return pd.DataFrame(data)


def dataframe_to_rows(df):
    """Convert edited DataFrame back to list of tuples."""
    rows = []
    for _, row in df.iterrows():
        vals = []
        for col_letter in COL_LETTERS:
            vals.append(row.get(col_letter, ""))
        rows.append(tuple(vals))
    return rows


def generate_excel_all_sheets(sheet_data_dict, original_file_bytes, all_sheet_names):
    """Generate an Excel file with all sheets, applying edits and highlighting edited cells."""
    wb = openpyxl.Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    edit_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    for sname in all_sheet_names:
        ws = wb.create_sheet(title=sname)
        # Header row
        for i, col_letter in enumerate(COL_LETTERS):
            ws.cell(row=1, column=i + 1, value=COLUMNS[col_letter])

        if sname in sheet_data_dict:
            sdata = sheet_data_dict[sname]
            rows = sdata.get("rows", [])
            edits = sdata.get("edits", set())
        else:
            # Sheet not yet validated — read original data
            rows = read_excel_rows(original_file_bytes, sheet_name=sname)
            edits = set()

        for row_i, row in enumerate(rows):
            for col_i, col_letter in enumerate(COL_LETTERS):
                idx = col_index(col_letter)
                val = row[idx] if idx < len(row) else ""
                cell = ws.cell(row=row_i + 2, column=col_i + 1, value=val)
                if (row_i, col_letter) in edits:
                    cell.fill = edit_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def run_validation(rows, checked_cols):
    results = []
    for row_i, row in enumerate(rows):
        row_result = {}
        for col_letter in COL_LETTERS:
            idx = col_index(col_letter)
            raw = row[idx] if idx < len(row) else None
            if col_letter not in checked_cols:
                row_result[col_letter] = {"value": str(raw).strip() if raw else "", "error": None, "skipped": True}
            else:
                display, error = validate_cell(col_letter, raw)
                row_result[col_letter] = {"value": display, "error": error, "skipped": False}
        results.append(row_result)
    return results

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🎓 Academic Personnel\nValidator")
    st.markdown("---")
    st.markdown("**How to use:**")
    st.markdown("1. Upload a filled `.xlsx` file\n2. Choose which columns to validate\n3. Click **Run Validation**")
    st.markdown("---")
    uploaded = st.file_uploader("Upload filled Excel file", type=["xlsx"])
    st.markdown("---")
    st.markdown("<small style='color:#666'>Template: AcademicPersonel_update.xlsx</small>", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
st.markdown('<div class="app-title">Academic Personnel Validator</div>', unsafe_allow_html=True)
st.markdown('<div class="app-subtitle">Upload → Select Columns → Validate</div>', unsafe_allow_html=True)

if not uploaded:
    st.info("👈 Upload a filled Excel file from the sidebar to get started.")
    st.stop()

# ─── SHEET SELECTION ───
file_bytes = uploaded.read()
uploaded.seek(0)
sheet_names = get_sheet_names(file_bytes)

selected_sheet = None
if len(sheet_names) > 1:
    st.markdown('<div class="section-header">Select Sheet</div>', unsafe_allow_html=True)
    st.info(f"This workbook has **{len(sheet_names)}** sheets: {', '.join(sheet_names)}")
    selected_sheet = st.selectbox("Choose a sheet to validate", sheet_names)
else:
    selected_sheet = sheet_names[0]

# ─── COLUMN SELECTION ───
st.markdown('<div class="section-header">Select Columns to Validate</div>', unsafe_allow_html=True)

# Initialize checkbox widget keys
for c in COL_LETTERS:
    if f"chk_{c}" not in st.session_state:
        st.session_state[f"chk_{c}"] = True


def toggle_all():
    current = all(st.session_state[f"chk_{c}"] for c in COL_LETTERS)
    for c in COL_LETTERS:
        st.session_state[f"chk_{c}"] = not current


toggle_col, _ = st.columns([1, 5])
with toggle_col:
    all_selected = all(st.session_state[f"chk_{c}"] for c in COL_LETTERS)
    st.button("✗ Deselect All" if all_selected else "✓ Select All", on_click=toggle_all)

# Render checkboxes in 4-column grid
cols_per_row = 4
grid_cols = st.columns(cols_per_row)
for i, col_letter in enumerate(COL_LETTERS):
    with grid_cols[i % cols_per_row]:
        label = f"**{col_letter}** — {COLUMNS[col_letter]}"
        st.checkbox(label, key=f"chk_{col_letter}")

checked_cols = [c for c in COL_LETTERS if st.session_state[f"chk_{c}"]]

st.markdown(f"<small style='color:#888'>{len(checked_cols)} of {len(COL_LETTERS)} columns selected</small>", unsafe_allow_html=True)

run_btn = st.button("▶  Run Validation")

# ─── PER-SHEET STATE ───
if "sheet_data" not in st.session_state:
    st.session_state.sheet_data = {}
if "file_bytes" not in st.session_state:
    st.session_state.file_bytes = None

# Store file bytes for download across sheets
st.session_state.file_bytes = file_bytes

# Show progress for multi-sheet workbooks
if len(sheet_names) > 1:
    fixed_sheets = [s for s in sheet_names if s in st.session_state.sheet_data]
    if fixed_sheets:
        progress_tags = ""
        for s in sheet_names:
            if s in st.session_state.sheet_data:
                sdata = st.session_state.sheet_data[s]
                edit_count = len(sdata.get("edits", set()))
                if edit_count > 0:
                    progress_tags += f'<span style="background:#1a3a2a;color:#4ade80;padding:0.15rem 0.5rem;border-radius:4px;font-size:0.75rem;margin-right:0.4rem;">{s} ({edit_count} edits)</span>'
                else:
                    progress_tags += f'<span style="background:#1a2a3a;color:#60a5fa;padding:0.15rem 0.5rem;border-radius:4px;font-size:0.75rem;margin-right:0.4rem;">{s} (validated)</span>'
            elif s == selected_sheet:
                progress_tags += f'<span style="background:#3a3a1a;color:#f1c40f;padding:0.15rem 0.5rem;border-radius:4px;font-size:0.75rem;margin-right:0.4rem;">{s} (current)</span>'
            else:
                progress_tags += f'<span style="background:#1a1a22;color:#666;padding:0.15rem 0.5rem;border-radius:4px;font-size:0.75rem;margin-right:0.4rem;">{s}</span>'
        st.markdown(f'<div style="margin:0.5rem 0;">{progress_tags}</div>', unsafe_allow_html=True)

# Check if we have data for the current sheet
has_sheet_data = selected_sheet in st.session_state.sheet_data

if not run_btn and not has_sheet_data:
    st.stop()

# ─── READ + VALIDATE ───
if run_btn:
    try:
        rows = read_excel_rows(file_bytes, sheet_name=selected_sheet)
    except Exception as e:
        st.error(f"Could not read Excel file: {e}")
        st.stop()
    results = run_validation(rows, checked_cols)
    st.session_state.sheet_data[selected_sheet] = {
        "results": results,
        "rows": rows,
        "checked": checked_cols,
        "edits": st.session_state.sheet_data.get(selected_sheet, {}).get("edits", set()),
    }

current_sheet_data = st.session_state.sheet_data[selected_sheet]
results = current_sheet_data["results"]
checked_cols = current_sheet_data["checked"]

# ─── SUMMARY DASHBOARD ───
total_rows = len(results)
row_error_counts = []
col_error_counts = {c: 0 for c in checked_cols}

for row in results:
    errs = sum(1 for c in checked_cols if row[c]["error"])
    row_error_counts.append(errs)
    for c in checked_cols:
        if row[c]["error"]:
            col_error_counts[c] += 1

total_errors = sum(row_error_counts)
valid_rows = sum(1 for e in row_error_counts if e == 0)
invalid_rows = total_rows - valid_rows

st.markdown('<div class="section-header">Summary</div>', unsafe_allow_html=True)
m1, m2, m3, m4 = st.columns(4)
with m1:
    st.markdown(f'<div class="metric-card blue"><div class="metric-num">{total_rows}</div><div class="metric-label">Total Rows</div></div>', unsafe_allow_html=True)
with m2:
    st.markdown(f'<div class="metric-card green"><div class="metric-num">{valid_rows}</div><div class="metric-label">Valid Rows</div></div>', unsafe_allow_html=True)
with m3:
    st.markdown(f'<div class="metric-card red"><div class="metric-num">{invalid_rows}</div><div class="metric-label">Invalid Rows</div></div>', unsafe_allow_html=True)
with m4:
    st.markdown(f'<div class="metric-card yellow"><div class="metric-num">{total_errors}</div><div class="metric-label">Total Errors</div></div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# Errors per column bar chart
if total_errors > 0:
    chart_data = {COLUMNS[c]: col_error_counts[c] for c in checked_cols if col_error_counts[c] > 0}
    df_chart = pd.DataFrame({"Column": list(chart_data.keys()), "Errors": list(chart_data.values())})
    df_chart = df_chart.sort_values("Errors", ascending=False)
    st.markdown('<div class="section-header">Errors by Column</div>', unsafe_allow_html=True)
    st.bar_chart(df_chart.set_index("Column"), color="#e74c3c", height=250)

# ─── TRACK EDITS ───
edits = current_sheet_data.get("edits", set())

# ─── COLOR-CODED TABLE ───
st.markdown('<div class="section-header">Data Table</div>', unsafe_allow_html=True)

filter_errors = st.checkbox("Show only rows with errors", value=False)
display_results = [(i, r) for i, r in enumerate(results) if not filter_errors or row_error_counts[i] > 0]

if not display_results:
    st.success("🎉 All rows are valid!")
else:
    # Build HTML table
    header_cells = "<th>#</th>" + "".join(
        f'<th title="{COLUMNS[c]}">{c}</th>' if c in checked_cols else f'<th style="opacity:0.4">{c}</th>'
        for c in COL_LETTERS
    )

    rows_html = ""
    for orig_i, row in display_results:
        err_count = row_error_counts[orig_i]
        edited_count = sum(1 for c in COL_LETTERS if (orig_i, c) in edits)
        badge = ""
        if err_count:
            badge += f'<span class="err-badge">{err_count}</span>'
        if edited_count:
            badge += f'<span class="edit-badge">{edited_count} edited</span>'
        cells = f'<td><b>{orig_i + 1}</b>{badge}</td>'
        for c in COL_LETTERS:
            cell = row[c]
            is_edited = (orig_i, c) in edits
            if is_edited and not cell.get("error"):
                cls = "cell-edited-ok"
            elif is_edited:
                cls = "cell-edited"
            elif cell["skipped"]:
                cls = "cell-skip"
            elif cell["error"]:
                cls = "cell-err"
            else:
                cls = "cell-ok"
            val = cell["value"] if cell["value"] else "—"
            title = cell["error"] or ("Edited" if is_edited else "")
            cells += f'<td class="{cls}" title="{title}">{val}</td>'
        rows_html += f"<tr>{cells}</tr>"

    table_html = f"""
    <div style="overflow-x:auto; border:1px solid #2a2a3a; border-radius:8px;">
    <table class="val-table">
        <thead><tr>{header_cells}</tr></thead>
        <tbody>{rows_html}</tbody>
    </table>
    </div>
    """
    st.markdown(table_html, unsafe_allow_html=True)
    st.markdown(
        "<small style='color:#888'>💡 Hover over a red cell to see the error. "
        "Blue cells = edited. Column letters match the template.</small>",
        unsafe_allow_html=True,
    )

# ─── ERROR DETAILS ───
error_rows = [(i, r) for i, r in enumerate(results) if row_error_counts[i] > 0]

if error_rows:
    st.markdown('<div class="section-header">Error Details</div>', unsafe_allow_html=True)
    for orig_i, row in error_rows:
        errors_in_row = [(c, row[c]["error"]) for c in checked_cols if row[c]["error"]]
        with st.expander(f"Row {orig_i + 1}  —  {len(errors_in_row)} error(s)"):
            for col_letter, err_msg in errors_in_row:
                st.markdown(
                    f'<div class="err-item">'
                    f'<span class="err-col-tag">{col_letter}</span>'
                    f'<b>{COLUMNS[col_letter]}:</b> {err_msg}'
                    f'</div>',
                    unsafe_allow_html=True,
                )

# ─── EDIT & FIX ───
current_rows = current_sheet_data["rows"]
col_display = {c: f"{c} - {COLUMNS[c]}" for c in COL_LETTERS}

if not error_rows:
    st.markdown('<div class="section-header">Edit & Fix</div>', unsafe_allow_html=True)
    st.success("🎉 No errors to fix!")
else:
    st.markdown('<div class="section-header">Edit & Fix</div>', unsafe_allow_html=True)
    st.markdown(
        f"<small style='color:#f87171'><b>{len(error_rows)}</b> row(s) with errors. "
        f"Only error rows are shown below — fix the highlighted columns and click <b>Save & Re-validate</b>.</small>",
        unsafe_allow_html=True,
    )

    # Show per-row error guide above the editor
    for orig_i, row in error_rows:
        err_cols = [c for c in checked_cols if row[c]["error"]]
        err_tags = " ".join(
            f'<span class="err-col-tag">{c}</span>' for c in err_cols
        )
        st.markdown(
            f'<div style="margin:0.2rem 0;font-size:0.82rem;color:#f87171;">'
            f'<b>Row {orig_i + 1}</b> → fix: {err_tags}</div>',
            unsafe_allow_html=True,
        )

    # Build DataFrame with ONLY error rows
    error_row_indices = [i for i, _ in error_rows]
    df_full = rows_to_dataframe(current_rows)
    df_errors = df_full.loc[error_row_indices].copy()

    # Add a "Row #" column at the front for reference
    df_errors.insert(0, "Row", [i + 1 for i in error_row_indices])

    # Only show columns that have errors + the Row column
    error_col_set = set()
    for orig_i, row in error_rows:
        for c in checked_cols:
            if row[c]["error"]:
                error_col_set.add(c)
    # Always include columns with errors, plus show all so user has context
    show_cols = ["Row"] + list(COL_LETTERS)
    df_errors_display = df_errors[show_cols].copy()

    # Rename data columns for clarity
    rename_map = {c: col_display[c] for c in COL_LETTERS}
    rename_map["Row"] = "Row #"
    df_errors_display = df_errors_display.rename(columns=rename_map)

    # Build column config
    col_config = {
        "Row #": st.column_config.NumberColumn("Row #", disabled=True, width="small"),
    }
    for col_letter in COL_LETTERS:
        display_name = col_display[col_letter]
        is_error_col = col_letter in error_col_set
        if col_letter in DROPDOWNS:
            col_config[display_name] = st.column_config.SelectboxColumn(
                f"{'⚠ ' if is_error_col else ''}{display_name}",
                options=DROPDOWNS[col_letter],
                required=col_letter in REQUIRED,
            )
        elif col_letter == "E":
            col_config[display_name] = st.column_config.NumberColumn(
                f"{'⚠ ' if is_error_col else ''}{display_name}",
                min_value=18, max_value=80, step=1,
            )
        elif col_letter == "W":
            col_config[display_name] = st.column_config.NumberColumn(
                f"{'⚠ ' if is_error_col else ''}{display_name}",
                min_value=0, step=0.5,
            )
        else:
            col_config[display_name] = st.column_config.TextColumn(
                f"{'⚠ ' if is_error_col else ''}{display_name}",
            )

    # Highlight error columns — move them to front after Row #
    error_col_order = [col_display[c] for c in COL_LETTERS if c in error_col_set]
    non_error_col_order = [col_display[c] for c in COL_LETTERS if c not in error_col_set]
    final_col_order = ["Row #"] + error_col_order + non_error_col_order
    df_errors_display = df_errors_display[final_col_order]

    edited_df = st.data_editor(
        df_errors_display,
        column_config=col_config,
        use_container_width=True,
        num_rows="fixed",
        key="data_editor",
        column_order=final_col_order,
    )

    # ─── SAVE & RE-VALIDATE ───
    revalidate_btn = st.button("🔄  Save & Re-validate")

    if revalidate_btn:
        # Reverse column rename
        reverse_rename = {v: k for k, v in rename_map.items()}
        edited_df_raw = edited_df.rename(columns=reverse_rename)

        # Map edits back to full dataset
        original_df = rows_to_dataframe(current_rows)
        new_edits = set(edits)
        for df_idx, orig_i in enumerate(error_row_indices):
            for col_letter in COL_LETTERS:
                orig_val = str(original_df.at[orig_i, col_letter]).strip()
                new_val = str(edited_df_raw.iloc[df_idx][col_letter]).strip()
                if orig_val != new_val:
                    new_edits.add((orig_i, col_letter))
                    # Update the full dataset
                    original_df.at[orig_i, col_letter] = new_val
        # Convert full edited df back to rows
        new_rows = dataframe_to_rows(original_df)

        # Re-run validation
        new_results = run_validation(new_rows, checked_cols)

        # Save back to per-sheet state
        st.session_state.sheet_data[selected_sheet] = {
            "results": new_results,
            "rows": new_rows,
            "checked": checked_cols,
            "edits": new_edits,
        }
        st.rerun()

# ─── DOWNLOAD ───
st.markdown('<div class="section-header">Download</div>', unsafe_allow_html=True)

# Show edit summary across all sheets
total_edits_all = 0
edit_summary_parts = []
for sname in sheet_names:
    if sname in st.session_state.sheet_data:
        sdata = st.session_state.sheet_data[sname]
        s_edits = sdata.get("edits", set())
        if s_edits:
            total_edits_all += len(s_edits)
            edit_summary_parts.append(f"<b>{sname}</b>: {len(s_edits)} cell(s)")

if edit_summary_parts:
    st.markdown(
        f"<small style='color:#60a5fa'>✏️ Total edits — {' | '.join(edit_summary_parts)}</small>",
        unsafe_allow_html=True,
    )

if len(sheet_names) > 1:
    st.markdown(
        "<small style='color:#888'>Download includes <b>all sheets</b> — "
        "edited sheets use your fixes, unedited sheets keep original data.</small>",
        unsafe_allow_html=True,
    )

excel_buf = generate_excel_all_sheets(
    st.session_state.sheet_data,
    st.session_state.file_bytes,
    sheet_names,
)
st.download_button(
    label="📥  Download Corrected Excel",
    data=excel_buf,
    file_name="corrected_academic_personnel.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.xml",
)